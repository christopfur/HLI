import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# Define the file path and sheet name
file_path = r'C:\Users\cmccullough\Desktop\Revenue\salesforce_reports_2024-12-04.xlsx'
sheet_name = 'Salesforce Opportunity Data'

# Load the Excel file
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Ensure the date columns are in datetime format
df['Membership Start Date'] = pd.to_datetime(df['Membership Start Date'], format='%Y-%m-%d')
df['Membership End Date'] = pd.to_datetime(df['Membership End Date'], format='%Y-%m-%d')

# Define the years for which we want to calculate active membership
years = [2020, 2021, 2022, 2023, 2024]

# Function to calculate the number of active days for a given year
def calculate_active_days_for_year(row, year):
    start_of_year = pd.Timestamp(f'{year}-01-01')
    end_of_year = pd.Timestamp(f'{year}-12-31')

    # Calculate the intersection of the membership period with the current year
    membership_start = max(row['Membership Start Date'], start_of_year)
    membership_end = min(row['Membership End Date'], end_of_year)

    if membership_start <= membership_end:
        return (membership_end - membership_start).days + 1
    return 0

# Function to convert days to years and round to the nearest quarter
def round_to_quarter_year(days):
    # Convert days to years (assuming 365.25 days per year for leap years)
    years = days / 365.25
    # Round to the nearest quarter (0, 0.25, 0.5, 0.75, or 1)
    rounded_years = round(years * 4) / 4
    return rounded_years

# Initialize a result DataFrame
result = pd.DataFrame()

# Calculate active membership for each year and round to the nearest quarter
for year in years:
    result[year] = df.apply(lambda row: round_to_quarter_year(calculate_active_days_for_year(row, year)), axis=1)

# Add 'Account ID' to the result
result['Account ID'] = df['Account ID']

# Count the number of opportunities per Account ID
account_opportunity_counts = df.groupby('Account ID').size().reset_index(name='# of Account Opportunities')

# Group by 'Account ID' and sum the active membership for each year
result = result.groupby('Account ID')[years].sum().reset_index()

# Merge the opportunity counts into the result DataFrame
result = result.merge(account_opportunity_counts, on='Account ID', how='left')

# Identify gap years
def has_gap_year(row):
    non_zero_years = [year for year in years if row[year] > 0]
    if len(non_zero_years) <= 1:
        return False  # No gap if there's only one or no active year
    # Check for gaps between non-zero years
    for i in range(len(non_zero_years) - 1):
        if non_zero_years[i + 1] - non_zero_years[i] > 1:
            return True
    return False

# Apply the gap year logic
result['Gap Year'] = result.apply(has_gap_year, axis=1)

# Update cumulative logic for each year
for i in range(1, len(years)):
    result[years[i]] = result.apply(
        lambda row: sum([row[years[j]] for j in range(0, i + 1)]) if row[years[i]] > 0 else 0, axis=1
    )

# Save the result to a new Excel file
output_path = r'C:\Users\cmccullough\Desktop\Revenue\active_membership_with_gap_years_and_opportunities_cumulative.xlsx'
result.to_excel(output_path, index=False)

# Freeze the top row
wb = load_workbook(output_path)
ws = wb.active
ws.freeze_panes = ws['A2']  # Freeze the top row
wb.save(output_path)

print(f"Active membership with cumulative values saved to: {output_path}")
