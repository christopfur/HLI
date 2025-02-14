import pandas as pd
from openpyxl import load_workbook

# File paths
input_file = r'C:\Users\cmccullough\Desktop\Revenue\salesforce_reports_2025-01-30 v2.xlsx'
output_file = r'C:\Users\cmccullough\Desktop\Revenue\processed_salesforce_data_monthly_invoices_2024.xlsx'

# Columns to extract
columns_to_extract = [
    'Opportunity ID',
    'Account Name',
    'Membership Location',
    'Membership Start Date',
    'Opportunity Product Short Name(s)',
    'Total Payments Received'
]

# Read the Excel file
df = pd.read_excel(input_file, usecols=columns_to_extract)

# Convert 'Account Name' to uppercase
df['Account Name'] = df['Account Name'].str.upper()

# Create the 'Invoice Number' column by extracting last 5 digits of 'Opportunity ID'
df['Invoice Number'] = df['Opportunity ID'].astype(str).str[-5:]

# Reorder columns to insert 'Invoice Number' after 'Opportunity ID'
cols = df.columns.tolist()
cols.insert(1, cols.pop(-1))  # Move 'Invoice Number' right after 'Opportunity ID'
df = df[cols]

# Convert 'Membership Start Date' to datetime format for filtering and grouping
df['Membership Start Date'] = pd.to_datetime(df['Membership Start Date'], errors='coerce')

# Filter for memberships starting on or after January 1, 2024, and on or before December 31, 2024
start_date_filter = pd.Timestamp("2024-01-01")
end_date_filter = pd.Timestamp("2024-12-31")
df = df[(df['Membership Start Date'] >= start_date_filter) & (df['Membership Start Date'] <= end_date_filter)]

#Filter > 0
df = df[(df['Total Payments Received'] > 0)]
        
# Drop rows where 'Membership Start Date' couldn't be converted (optional)
df = df.dropna(subset=['Membership Start Date'])

# Convert 'Membership Start Date' to date only (removing time component)
df['Membership Start Date'] = df['Membership Start Date'].dt.date

# Group data by year and month
grouped_data = df.groupby([df['Membership Start Date'].apply(lambda x: x.year),
                           df['Membership Start Date'].apply(lambda x: x.month)])

# Save each group to a separate worksheet in a new Excel file
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for (year, month), data in grouped_data:
        sheet_name = f"{year}_{month:02d}"  # e.g., "2024_01" for January 2024
        data.to_excel(writer, sheet_name=sheet_name, index=False)

# Adjust column width based on content and freeze the first row
wb = load_workbook(output_file)

for sheet in wb.sheetnames:
    worksheet = wb[sheet]
    # Freeze the first row
    worksheet.freeze_panes = worksheet['A2']
    
    # Adjust column widths
    for column_cells in worksheet.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter  # Get the column letter
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # Add some padding to the width
        worksheet.column_dimensions[col_letter].width = adjusted_width

# Save the workbook with adjusted column widths and frozen header row
wb.save(output_file)

print("Processing complete. Data saved with adjusted column widths, frozen header rows, and uppercase 'Account Name' to:", output_file)