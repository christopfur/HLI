import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

# Define the folder path
folder_path = r'C:\Users\cmccullough\Desktop\Payroll Automation 1.0\Payrolls June & July 2024\output'

# Function to process each Excel file
def process_excel_file(file_path):
    try:
        # Read the Excel file
        xl = pd.ExcelFile(file_path)
        
        # Check if 'GL Costs 2' tab exists
        if 'GL Costs 2' in xl.sheet_names:
            # Read the 'GL Costs 2' tab into a DataFrame
            df = xl.parse('GL Costs 2')
            
            # Filter the DataFrame to match the conditions
            filtered_df = df.groupby(['GL Category Code', 'GL Category', 'GL Code']).agg({
                'Total Dollar Amount': 'sum'
            }).reset_index()
            
            return filtered_df
        else:
            print(f"Warning: 'GL Costs 2' tab not found in {file_path}")
            return None
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return None

# Create a new Excel workbook
wb = Workbook()
# Remove the default sheet created
wb.remove(wb.active)

# List to store individual DataFrames for summary
summary_df_list = []

# Process all Excel files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        result_df = process_excel_file(file_path)
        if result_df is not None:
            # Truncate the filename to the first 30 characters for the sheet name
            truncated_filename = filename[:30]
            
            # Add the DataFrame to the summary list
            summary_df_list.append(result_df)
            
            # Add a new sheet with the truncated filename as the sheet name
            ws = wb.create_sheet(title=truncated_filename)
            
            # Convert DataFrame to rows and append to worksheet
            for r in dataframe_to_rows(result_df, index=False, header=True):
                ws.append(r)
            
            # Lock header row
            ws.freeze_panes = 'A2'
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

# Concatenate all summary DataFrames
if summary_df_list:
    summary_df = pd.concat(summary_df_list, ignore_index=True)
    
    # Aggregate the total sum
    total_sum = summary_df['Total Dollar Amount'].sum()
    
    # Group by first three columns
    grouped_df = summary_df.groupby(['GL Category Code', 'GL Category', 'GL Code']).agg({
        'Total Dollar Amount': 'sum'
    }).reset_index()
    
    # Add the summary DataFrame as a new sheet at the beginning
    summary_ws = wb.create_sheet(title="Summary", index=0)
    
    # Convert DataFrame to rows and append to worksheet
    for r in dataframe_to_rows(grouped_df, index=False, header=True):
        summary_ws.append(r)
    
    # Lock header row
    summary_ws.freeze_panes = 'A2'
    
    # Adjust column widths
    for col in summary_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        summary_ws.column_dimensions[column].width = adjusted_width

    # Bold the first title row
    for cell in summary_ws["1:1"]:
        cell.font = Font(bold=True)
    
    # Color the Summary tab red
    summary_ws.sheet_properties.tabColor = "FF0000"

# Save results to Excel
output_folder = r'C:\Users\cmccullough\Desktop\Payroll Automation 1.0\output'
today_date = datetime.today().strftime('%Y-%m-%d')
output_file = os.path.join(output_folder, f'summarized_results_{today_date}.xlsx')
wb.save(output_file)
print(f"Results saved to: {output_file}")
